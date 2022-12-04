#IMPORTS
import os
from customer import Customer
from datetime import datetime
from openpyxl import load_workbook #excel api
from openpyxl.utils import get_column_letter
#from openpyxl.styles import Style, Font
import openpyxl
from win32com import client

#Global constants
NORMAL = '\033[0;37;40m'
NORMAL_BOLD = '\033[2;30;47m'
GREEN = '\033[0;32;40m'
GREEN_BOLD = '\033[0;30;42m'
PROMPT = f"{NORMAL}>> "
ERROR = '\033[0;31;40m'
BLUE = '\033[1;34;40m'
YELLOW = '\033[1;33;40m'

CURRENT_DIR = os.getcwd()



class InvoiceManager:
    def __init__(self, invoice_data_file_path, invoice_template_file_path, invoices_file_path):
        self.INVOICE_DATA_FILE_PATH = invoice_data_file_path
        self.INVOICE_TEMPLATE_FILE_PATH = invoice_template_file_path
        self.INVOICES_FILE_PATH = invoices_file_path
        self.work_book_data = None
        self.work_sheet_data = None
        self.work_book_template = None
        self.work_sheet_template = None
        self.customer = None
        self.customers = []
        self.specific_accounts = []
        self.ommited_accounts = []
        self.month = None
        self.year = None

        #Testing if the files exist (Invoice data file)
        try:
            self.INVOICE_DATA_FILE = open(self.INVOICE_DATA_FILE_PATH)
        except FileNotFoundError:
            print(f"{ERROR}No invoice data file was found. The program will now be stopped.")
            quit()
        finally:
            print(f"{GREEN}Sucessfully loaded the invoice data file")

        #Testing if the files exist (Invoice template file)
        try:
            self.INVOICE_TEMPLATE_FILE = open(self.INVOICE_TEMPLATE_FILE_PATH)
        except FileNotFoundError:
            print(f"{ERROR}No invoice template file was found. The program will now be stopped.")
            quit()
        finally:
            print(f"{GREEN}Sucessfully loaded the invoice template file")

    def set_up_workbooks(self):
        #Setting up template sheet

        self.work_book_template = load_workbook(self.INVOICE_TEMPLATE_FILE_PATH)
        self.work_sheet_template = self.work_book_template.active

    def set_year_path(self, year):
        #Sets the year path as well as updates the invoices path
        if not os.path.exists(self.INVOICES_FILE_PATH + year):
            os.mkdir(self.INVOICES_FILE_PATH + year + '\\')
        self.INVOICES_FILE_PATH += year + '\\'
        self.year = year

    def set_month_path(self, month):
        #Sets the month path as well as updates the invoices path
        if not os.path.exists(self.INVOICES_FILE_PATH + month):
            os.mkdir(self.INVOICES_FILE_PATH + month + '\\')
        self.INVOICES_FILE_PATH += month + '\\'
        self.month = month

    def get_data_sheet(self):
        #Returns the sheet
        not_found = True

        while not_found:
            print(f"\n{NORMAL_BOLD}Which sheet would you like to access")
            sheet = input(PROMPT)

            #Looks for the sheet in the work_book
            for s in range(len(self.work_book_data.sheetnames)):
                if self.work_book_data.sheetnames[s] == sheet: #IF FOUND
                    self.work_book_data.active = s
                    print(f"{BLUE}Sheet: {sheet} is selected.")
                    return self.work_book_data[sheet]
            print(f"{ERROR}Cannot find the {sheet} sheet.")

    def set_data_sheet(self, sheet):
        #Assigns the sheet to be used
        self.work_sheet_data = sheet

    def fill_in_template(self, customer):
        #Completes the template with the customer's details
        #DATE
        self.work_sheet_template['D3'].value = str(self.month) + '-' + str(self.year)
        #ACCOUNT NUMBER
        self.work_sheet_template['B4'].value = 'Acc no: ' + str(customer.acc_num)
        #BBF AMOUNT
        self.work_sheet_template['F8'].value = customer.bbf_amount
        #self.work_sheet_template['F8'].style = Style(number_format='R#.##')
        #BBF DATE
        self.work_sheet_template['D8'].value = customer.bbf_date
        #FEE AMOUNT
        self.work_sheet_template['F9'].value = customer.fee_amount
        #self.work_sheet_template['F9'].style = Style(number_format='R#.##')
        #INVOICE ISSUE DATE
        self.work_sheet_template['G4'].value = customer.invoice_issue_date
        #FEE DATE
        self.work_sheet_template['D9'].value = customer.fee_date
        #ADJUSTMENTS
        self.work_sheet_template['G12'].value = customer.adjustments
        #NAME
        self.work_sheet_template['B6'].value = customer.name
        #ADDRESS
        self.work_sheet_template['B7'].value = customer.address
        #SUBURB
        self.work_sheet_template['B8'].value = customer.suburb
        #POSTAL CODE
        self.work_sheet_template['B9'].value = customer.postal
        #INVOICE NUMBER
        self.work_sheet_template['E4'].value = customer.invoice_num
        #REFERENCE NUMBER
        self.work_sheet_template['B18'].value += ' ' + customer.reference

        #Adding image
        logo = openpyxl.drawing.image.Image(CURRENT_DIR + "\\invoice_generation\\logo.png")
        logo.anchor = 'B1'
        self.work_sheet_template.add_image(logo)

        #Printing successfuly message
        print(f'{GREEN}Invoice template modification successful')

        #Saving the file
        return self.save_file_as_xlsx(self.work_book_template, customer)

    def save_file_as_xlsx(self, file, customer):
        #Preparing the file to be converted to pdf by saving it first
        file_name = str(customer.acc_num) + '.xlsx'
        file.save(self.INVOICES_FILE_PATH + file_name)
        file.close()
        print(f'{GREEN}File {file_name} successfully saved')
        return file_name

    def convert_to_pdf(self, customer, file_name):
        #Converting the file to pdf
        print(f'{YELLOW}Converting {str(customer.acc_num)}.xlsx to PDF format')
        excel = client.Dispatch('Excel.Application')
        temp_sheets = excel.Workbooks.Open(self.INVOICES_FILE_PATH + file_name)
        work_sheets = temp_sheets.Worksheets[0]
        work_sheets.Visible = 1
        work_sheets.ExportAsFixedFormat(0, self.INVOICES_FILE_PATH + str(customer.acc_num) + '.pdf')
        temp_sheets.Close(True)
        print(f'{GREEN}File {file_name}.pdf successfully created')

        #Removing temporary excel file
        print(f'{BLUE}Removing temporary excel file')
        os.remove(self.INVOICES_FILE_PATH + file_name)
        print(f'{GREEN}Excel file removal successful')
        print(f'{GREEN_BOLD}Invoice successfully created for Account number {customer.acc_num}{NORMAL}')

    def populate_customer_list(self, acc_num, bbf_amount, fee_amount, less_credit, adjustments, bbf_date, fee_date, invoice_issue_date, name, address, suburb, postal, invoice_num, reference):
        self.customer = Customer(acc_num,
                                 bbf_amount,
                                 fee_amount,
                                 less_credit,
                                 adjustments,
                                 bbf_date,
                                 fee_date,
                                 invoice_issue_date,
                                 name,
                                 address,
                                 suburb,
                                 postal,
                                 invoice_num,
                                 reference)
        self.customers.append(self.customer)
        return self.customers

    def generate_invoices(self, list):
        #Generating the templates for each customer in the list
        for i in range(len(list)):
            self.convert_to_pdf(list[i], self.fill_in_template(list[i]))
